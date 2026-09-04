"""Three-line (booktabs) xlsx writer.

Layout produced, which is exactly what the `tables_threeline` gate verifies:

    row 1            title, unruled
    row 2            header, rule above and rule below
    rows 3..n        data, unruled
    row n            bottom rule
    rows n+1..       footnotes, unruled

No vertical rules, no interior horizontal rules. Use this rather than styling cells
by hand, so every table in the paper is identical.
"""
from __future__ import annotations

import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required for table building.\n"
        "  uv pip install --python .venv/Scripts/python.exe openpyxl"
    ) from exc

RULE = Side(style="thin", color="FF000000")
BASE_FONT = "Arial"
TITLE_PT = 10
BODY_PT = 9
FOOT_PT = 8
MAX_CELL_CHARS = 300
MAX_FOOTNOTE_CHARS = 1500


class TableSpec(dict):
    """{sheet, title, header, rows, footnotes, align, col_widths}"""


def _check(spec: dict) -> None:
    for key in ("sheet", "title", "header", "rows"):
        if not spec.get(key):
            raise ValueError(f"table spec missing '{key}'")
    if not spec.get("footnotes"):
        raise ValueError(
            f"'{spec['sheet']}': footnotes are required. A three-line table needs at least "
            "an abbreviations or data-format note."
        )
    ncol = len(spec["header"])
    for i, row in enumerate(spec["rows"], 1):
        if len(row) != ncol:
            raise ValueError(
                f"'{spec['sheet']}' row {i} has {len(row)} cells but the header has {ncol}"
            )
        for cell in row:
            if isinstance(cell, str) and len(cell) > MAX_CELL_CHARS:
                raise ValueError(
                    f"'{spec['sheet']}' row {i}: cell of {len(cell)} chars. A table is not a "
                    "place for prose - move it to the footnote or the manuscript."
                )
    total = sum(len(str(f)) for f in spec["footnotes"])
    if total > MAX_FOOTNOTE_CHARS:
        raise ValueError(
            f"'{spec['sheet']}': footnotes total {total} chars (limit {MAX_FOOTNOTE_CHARS}). "
            "That is an analysis report, not a table footnote."
        )
    if not re.match(r"^\s*(table|supplementary table)\s+S?\d+[.:]", spec["title"], re.I):
        raise ValueError(
            f"'{spec['sheet']}': title must start with e.g. 'Table 1.' or 'Table S1.' - got "
            f"{spec['title'][:40]!r}"
        )


def _autowidth(header: list, rows: list, explicit: list | None) -> list[float]:
    if explicit:
        return list(explicit)
    widths = []
    for i, head in enumerate(header):
        longest = max([len(str(head))] + [len(str(r[i])) for r in rows] or [0])
        widths.append(min(48.0, max(10.0, longest * 1.15 + 2)))
    return widths


def _write_sheet(ws, spec: dict) -> None:
    header = spec["header"]
    rows = spec["rows"]
    footnotes = spec["footnotes"]
    ncol = len(header)
    align = spec.get("align") or ["left"] + ["center"] * (ncol - 1)
    last_col = get_column_letter(ncol)

    # --- title (row 1, unruled) ---------------------------------------
    ws.cell(row=1, column=1, value=spec["title"])
    ws.cell(row=1, column=1).font = Font(name=BASE_FONT, size=TITLE_PT, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    if ncol > 1:
        ws.merge_cells(f"A1:{last_col}1")

    # --- header (row 2, ruled above and below) -------------------------
    for c, text in enumerate(header, 1):
        cell = ws.cell(row=2, column=c, value=text)
        cell.font = Font(name=BASE_FONT, size=BODY_PT, bold=True)
        cell.alignment = Alignment(horizontal=align[c - 1], vertical="bottom", wrap_text=True)
        cell.border = Border(top=RULE, bottom=RULE)

    # --- data ----------------------------------------------------------
    first_data = 3
    for r, row in enumerate(rows, first_data):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(name=BASE_FONT, size=BODY_PT)
            cell.alignment = Alignment(horizontal=align[c - 1], vertical="top", wrap_text=False)
    last_data = first_data + len(rows) - 1

    # --- bottom rule ---------------------------------------------------
    for c in range(1, ncol + 1):
        ws.cell(row=last_data, column=c).border = Border(bottom=RULE)

    # --- footnotes (unruled) -------------------------------------------
    for i, note in enumerate(footnotes):
        r = last_data + 1 + i
        cell = ws.cell(row=r, column=1, value=str(note))
        cell.font = Font(name=BASE_FONT, size=FOOT_PT)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        if ncol > 1:
            ws.merge_cells(f"A{r}:{last_col}{r}")

    for i, w in enumerate(_autowidth(header, rows, spec.get("col_widths")), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{first_data}"


def write_workbook(path: str | Path, tables: list[dict]) -> Path:
    """One workbook, one sheet per table. Use this for the supplementary tables."""
    if not tables:
        raise ValueError("no tables given")
    for spec in tables:
        _check(spec)
    seen = [t["sheet"] for t in tables]
    if len(seen) != len(set(seen)):
        raise ValueError(f"duplicate sheet names: {seen}")
    wb = Workbook()
    wb.remove(wb.active)
    for spec in tables:
        ws = wb.create_sheet(title=spec["sheet"][:31])
        _write_sheet(ws, spec)
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    print(f"wrote {p}  ({len(tables)} sheet(s))")
    return p


def write_table(path: str | Path, sheet: str, title: str, header: list,
                rows: list, footnotes: list, align: list | None = None,
                col_widths: list | None = None) -> Path:
    """One workbook holding one table. Use this for each main table."""
    return write_workbook(path, [{
        "sheet": sheet, "title": title, "header": header, "rows": rows,
        "footnotes": footnotes, "align": align, "col_widths": col_widths,
    }])


def fmt(value, decimals: int = 2, pct: bool = False) -> str:
    """Consistent numeric rendering. Keeps trailing zeros, which reviewers expect."""
    if value is None or value == "":
        return "-"
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{f:.{decimals}f}%" if pct else f"{f:.{decimals}f}"


def p_value(p) -> str:
    """APA/medical convention: report small p as an inequality, never as 0.000."""
    try:
        f = float(p)
    except (TypeError, ValueError):
        return str(p)
    if f < 0.001:
        return "<0.001"
    return f"{f:.3f}" if f < 0.01 else f"{f:.2f}"


def ci(estimate, low, high, decimals: int = 2) -> str:
    return f"{fmt(estimate, decimals)} ({fmt(low, decimals)} to {fmt(high, decimals)})"
